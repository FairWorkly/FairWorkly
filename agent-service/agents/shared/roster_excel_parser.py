"""
Roster Excel Parser

Parses roster Excel files (.xlsx) for shift data import.

This module handles:
- Reading Excel files with roster/shift data
- Validating required columns and data formats
- Converting Excel rows to typed RosterEntry objects
- Error handling with row-level error reporting
- Warning generation for data quality issues
"""

from typing import Any, Optional
from datetime import date, time, datetime, timedelta
from pathlib import Path
from enum import Enum
from decimal import Decimal
import re
import unicodedata

from fastapi import UploadFile
from pydantic import BaseModel, ConfigDict, EmailStr, TypeAdapter, ValidationError, computed_field
from pydantic.alias_generators import to_camel
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


class ParseIssueSeverity(Enum):
    """Severity level for parse issues."""
    ERROR = "error"
    WARNING = "warning"


class ParseIssue(BaseModel):
    """Represents a parsing issue (error or warning)."""
    model_config = ConfigDict(
        alias_generator=to_camel,
        populate_by_name=True,  # Allow both snake_case and camelCase
        use_enum_values=True,  # Serialize enum as value string
    )

    severity: ParseIssueSeverity
    code: str
    message: str
    row: int  # Real Excel row number
    column: Optional[str] = None
    value: Optional[str] = None


class ParseIssueError(Exception):
    """Exception wrapper for row-level parse issues."""
    def __init__(self, issue: ParseIssue):
        super().__init__(issue.message)
        self.issue = issue


class RosterEntry(BaseModel):
    """Represents a single roster entry (shift) parsed from Excel."""
    model_config = ConfigDict(
        alias_generator=to_camel,
        populate_by_name=True,  # Allow both snake_case and camelCase
    )

    excel_row: int  # Real Excel row number for traceability
    employee_email: EmailStr  # Required: Primary matching key
    employee_number: Optional[str] = None  # Optional: Secondary matching key
    employee_name: Optional[str] = None  # Optional: Display only
    date: date
    start_time: time
    end_time: time
    is_overnight: bool = False  # True if end_time < start_time (crosses midnight)
    has_meal_break: bool = False
    meal_break_duration: Optional[int] = None
    has_rest_breaks: bool = False
    rest_breaks_duration: Optional[int] = None
    is_public_holiday: bool = False
    public_holiday_name: Optional[str] = None
    is_on_call: bool = False
    location: Optional[str] = None
    notes: Optional[str] = None

    @computed_field
    @property
    def duration_hours(self) -> Decimal:
        """
        Shift duration in hours (gross, before breaks).
        Handles overnight shifts (e.g., 22:00 - 06:00 = 8 hours).
        """
        start_dt = datetime.combine(self.date, self.start_time)
        end_dt = datetime.combine(self.date, self.end_time)
        if self.is_overnight:
            end_dt += timedelta(days=1)
        duration = end_dt - start_dt
        return Decimal(str(duration.total_seconds() / 3600)).quantize(Decimal("0.01"))

    @computed_field
    @property
    def net_hours(self) -> Decimal:
        """
        Net working hours (duration minus breaks).
        """
        total_break_minutes = (self.meal_break_duration or 0) + (self.rest_breaks_duration or 0)
        break_hours = Decimal(str(total_break_minutes / 60))
        return (self.duration_hours - break_hours).quantize(Decimal("0.01"))


class EmployeeEntry(BaseModel):
    """Represents a single employee entry parsed from Excel."""
    model_config = ConfigDict(
        alias_generator=to_camel,
        populate_by_name=True,  # Allow both snake_case and camelCase
    )

    excel_row: int
    name: str
    email: str
    role: str
    department: Optional[str] = None
    start_date: Optional[date] = None


class RosterParseResult(BaseModel):
    """
    Result of parsing a roster Excel file.

    Contains parsed entries and summary statistics.
    """
    model_config = ConfigDict(
        alias_generator=to_camel,
        populate_by_name=True,
    )

    entries: list[RosterEntry]

    @computed_field
    @property
    def week_start_date(self) -> Optional[date]:
        """Earliest shift date in the roster."""
        if not self.entries:
            return None
        return min(entry.date for entry in self.entries)

    @computed_field
    @property
    def week_end_date(self) -> Optional[date]:
        """Latest shift date in the roster."""
        if not self.entries:
            return None
        return max(entry.date for entry in self.entries)

    @computed_field
    @property
    def total_shifts(self) -> int:
        """Total number of shifts parsed."""
        return len(self.entries)

    @computed_field
    @property
    def total_hours(self) -> Decimal:
        """Total gross hours across all shifts."""
        if not self.entries:
            return Decimal("0.00")
        total = sum((entry.duration_hours for entry in self.entries), Decimal("0.00"))
        return total.quantize(Decimal("0.01"))

    @computed_field
    @property
    def unique_employees(self) -> int:
        """Number of unique employees in the roster."""
        if not self.entries:
            return 0
        return len({str(entry.employee_email).lower() for entry in self.entries})


class RosterExcelParser:
    """
    Parses roster Excel files into typed shift entries.

    Usage:
        parser = RosterExcelParser()

        # Parse roster Excel file
        result, issues = parser.parse_roster_excel("roster.xlsx")
        for entry in result.entries:
            print(f"{entry.employee_email}: {entry.date} {entry.start_time}-{entry.end_time}")

        # Or read raw Excel data
        data = parser.read_excel("roster.xlsx")
    """

    # Header aliases: canonical_key -> list of accepted variations
    # Note: Roster uses "employee_email" prefix, Employee uses "email"
    # This avoids alias conflicts between the two file types
    HEADER_ALIASES: dict[str, list[str]] = {
        # Roster-specific fields (use "employee_" prefix)
        "employee_email": ["employee email", "employee_email", "员工邮箱", "staff email", "worker email"],
        "employee_number": ["employee number", "employee_number", "emp number", "emp_number", "employee id", "employee_id", "员工编号", "staff number"],
        "employee_name": ["employee name", "employee_name", "员工姓名", "staff name", "worker name"],
        "date": ["date", "shift date", "shift_date", "日期"],
        "start_time": ["start time", "start_time", "start", "开始时间"],
        "end_time": ["end time", "end_time", "end", "finish time", "finish_time", "结束时间"],
        "has_meal_break": ["has meal break", "has_meal_break", "meal break", "meal_break"],
        "meal_break_duration": ["meal break duration", "meal_break_duration", "meal break mins", "meal_break_mins"],
        "has_rest_breaks": ["has rest breaks", "has_rest_breaks", "rest breaks", "rest_breaks"],
        "rest_breaks_duration": ["rest breaks duration", "rest_breaks_duration", "rest break mins", "rest_break_mins"],
        "is_public_holiday": ["is public holiday", "is_public_holiday", "public holiday", "public_holiday"],
        "public_holiday_name": ["public holiday name", "public_holiday_name", "holiday name", "holiday_name"],
        "is_on_call": ["is on call", "is_on_call", "on call", "on_call"],
        "location": ["location", "work location", "work_location", "地点"],
        "notes": ["notes", "note", "comments", "comment", "备注"],
        # Employee-specific fields (short names)
        "name": ["name", "姓名", "full name"],
        "email": ["email", "e-mail", "邮箱", "mail"],
        "role": ["role", "position", "title", "职位", "job title"],
        "department": ["department", "dept", "部门"],
        "start_date": ["start date", "start_date", "hire date", "hire_date", "入职日期"],
    }

    # Required canonical keys for roster files
    ROSTER_REQUIRED_KEYS = ["employee_email", "date", "start_time", "end_time"]

    # Required canonical keys for employee files
    EMPLOYEE_REQUIRED_KEYS = ["name", "email", "role"]

    # Reusable email validator
    _email_validator = TypeAdapter(EmailStr)

    def __init__(self):
        # Build reverse lookup: normalized alias -> canonical key
        self._alias_to_canonical: dict[str, str] = {}
        for canonical, aliases in self.HEADER_ALIASES.items():
            for alias in aliases:
                self._alias_to_canonical[self._normalize_header(alias)] = canonical

    @staticmethod
    def _normalize_header(header: str) -> str:
        """
        Normalize header string for robust matching.

        Handles:
        - Multiple spaces ("Employee  Email" -> "employee email")
        - Newlines ("Employee\\nEmail" -> "employee email")
        - Leading/trailing whitespace
        - Full-width characters (全角 -> 半角)
        - Special characters like *, #, etc.
        """
        if not header:
            return ""
        # Normalize unicode (full-width -> half-width, etc.)
        header = unicodedata.normalize("NFKC", header)
        # Lowercase
        header = header.lower()
        # Replace any whitespace (including newlines, tabs) with single space
        header = re.sub(r"\s+", " ", header)
        # Remove common noise characters
        header = re.sub(r"[*#\-_.:：]", " ", header)
        # Collapse multiple spaces again after removing noise
        header = re.sub(r"\s+", " ", header)
        # Strip
        return header.strip()

    def read_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None
    ) -> list[dict[str, Any]]:
        """
        Read an Excel file and return data as a list of dictionaries.

        Each dict includes '__row__' key with the real Excel row number.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to read (default: first sheet)

        Returns:
            List of dictionaries, where each dict represents a row with
            column headers as keys and '__row__' for the real row number.

        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the file format is invalid or corrupted
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"Invalid file format. Expected .xlsx, got: {path.suffix}")

        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to read Excel file: {str(e)}")

        # Select sheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # Stream rows instead of loading all into memory
        row_iter = sheet.iter_rows()

        # Read header row first
        try:
            header_row = next(row_iter)
        except StopIteration:
            workbook.close()
            return []

        headers = [self._get_cell_value(cell) for cell in header_row]
        headers = [h.strip() if isinstance(h, str) else h for h in headers]

        # Stream remaining rows
        data = []
        for row in row_iter:
            row_dict: dict[str, Any] = {}
            is_empty_row = True
            real_row_num: Optional[int] = None

            # Get real Excel row number from any cell that has it
            for cell in row:
                if hasattr(cell, 'row') and cell.row is not None:
                    real_row_num = cell.row
                    break

            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    value = self._get_cell_value(cell)
                    if value is not None:
                        is_empty_row = False
                    row_dict[headers[i]] = value

            # Skip completely empty rows
            if not is_empty_row and real_row_num is not None:
                row_dict["__row__"] = real_row_num
                data.append(row_dict)

        workbook.close()
        return data

    def _normalize_headers(self, row: dict[str, Any]) -> dict[str, Any]:
        """
        Normalize header names to canonical keys using alias mapping.

        Returns a new dict with canonical keys.
        """
        normalized: dict[str, Any] = {}
        for key, value in row.items():
            if key == "__row__":
                normalized[key] = value
                continue
            # Normalize the key for matching
            key_normalized = self._normalize_header(key)
            canonical = self._alias_to_canonical.get(key_normalized)
            if canonical:
                normalized[canonical] = value
            else:
                # Keep normalized key for unknown headers
                normalized[key_normalized] = value
        return normalized

    def validate_headers(
        self,
        data: list[dict[str, Any]],
        required_keys: list[str]
    ) -> tuple[bool, list[str]]:
        """
        Validate that required headers exist in the data (using alias matching).

        Args:
            data: List of dictionaries from read_excel
            required_keys: List of required canonical key names

        Returns:
            Tuple of (is_valid, missing_keys)
        """
        if not data:
            return False, required_keys

        # Normalize first row to get canonical keys
        normalized = self._normalize_headers(data[0])
        # Compare case-insensitively to avoid header casing issues
        available_keys = {key.lower() for key in normalized.keys()}

        missing = []
        for required in required_keys:
            if required.lower() not in available_keys:
                missing.append(required)

        return len(missing) == 0, missing

    def parse_roster_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None
    ) -> tuple[RosterParseResult, list[ParseIssue]]:
        """
        Parse a roster Excel file into RosterEntry objects.

        Args:
            file_path: Path to the roster Excel file
            sheet_name: Name of the sheet to read (default: first sheet)

        Returns:
            Tuple of (result, issues)
            - result: RosterParseResult with entries + summary
            - issues: List of ParseIssue (errors and warnings)

        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the file format is invalid
        """
        data = self.read_excel(file_path, sheet_name)

        if not data:
            return RosterParseResult(entries=[]), [ParseIssue(
                row=0,
                severity=ParseIssueSeverity.ERROR,
                code="EMPTY_FILE",
                message="Excel file is empty or contains no data rows"
            )]

        # Validate required headers
        is_valid, missing = self.validate_headers(data, self.ROSTER_REQUIRED_KEYS)
        if not is_valid:
            return RosterParseResult(entries=[]), [ParseIssue(
                row=0,
                severity=ParseIssueSeverity.ERROR,
                code="MISSING_REQUIRED_COLUMNS",
                message=f"Missing required columns: {', '.join(missing)}"
            )]

        entries: list[RosterEntry] = []
        issues: list[ParseIssue] = []

        for row in data:
            real_row = row.get("__row__", 0)
            try:
                entry, row_warnings = self._parse_roster_row(row, real_row)
                entries.append(entry)
                issues.extend(row_warnings)
            except ParseIssueError as e:
                issues.append(e.issue)
            except Exception as e:
                issues.append(ParseIssue(
                    row=real_row,
                    severity=ParseIssueSeverity.ERROR,
                    code="ROW_PARSE_ERROR",
                    message=str(e)
                ))

        return RosterParseResult(entries=entries), issues

    def parse_employee_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None
    ) -> tuple[list[EmployeeEntry], list[ParseIssue]]:
        """
        Parse an employee Excel file into EmployeeEntry objects.

        Args:
            file_path: Path to the employee Excel file
            sheet_name: Name of the sheet to read (default: first sheet)

        Returns:
            Tuple of (list of valid EmployeeEntry objects, list of ParseIssue errors)
        """
        data = self.read_excel(file_path, sheet_name)

        if not data:
            return [], [ParseIssue(
                row=0,
                severity=ParseIssueSeverity.ERROR,
                code="EMPTY_FILE",
                message="Excel file is empty or contains no data rows"
            )]

        # Validate required headers
        is_valid, missing = self.validate_headers(data, self.EMPLOYEE_REQUIRED_KEYS)
        if not is_valid:
            return [], [ParseIssue(
                row=0,
                severity=ParseIssueSeverity.ERROR,
                code="MISSING_REQUIRED_COLUMNS",
                message=f"Missing required columns: {', '.join(missing)}"
            )]

        entries: list[EmployeeEntry] = []
        errors: list[ParseIssue] = []

        for row in data:
            real_row = row.get("__row__", 0)
            try:
                entry = self._parse_employee_row(row, real_row)
                entries.append(entry)
            except ParseIssueError as e:
                errors.append(e.issue)
            except Exception as e:
                errors.append(ParseIssue(
                    row=real_row,
                    severity=ParseIssueSeverity.ERROR,
                    code="ROW_PARSE_ERROR",
                    message=str(e)
                ))

        return entries, errors

    async def parse_excel(self, file: UploadFile) -> list[dict[str, Any]]:
        """
        Parse an uploaded Excel file.

        Args:
            file: FastAPI UploadFile object

        Returns:
            List of dictionaries representing the Excel data
        """
        import tempfile
        import os

        # Save uploaded file to temp location
        suffix = Path(file.filename).suffix if file.filename else ".xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name

        try:
            return self.read_excel(tmp_path)
        finally:
            os.unlink(tmp_path)

    def _parse_roster_row(self, row: dict[str, Any], row_num: int) -> tuple[RosterEntry, list[ParseIssue]]:
        """Parse a single row into a RosterEntry with warnings."""
        warnings: list[ParseIssue] = []

        # Normalize headers to canonical keys
        normalized = self._normalize_headers(row)

        # Required fields
        raw_email = normalized.get("employee_email")
        employee_email = self._get_string(raw_email)
        if not employee_email:
            raise ParseIssueError(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.ERROR,
                code="MISSING_REQUIRED_FIELD",
                message="Employee Email is required",
                column="employee_email"
            ))
        try:
            self._email_validator.validate_python(employee_email)
        except ValidationError:
            raise ParseIssueError(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.ERROR,
                code="INVALID_EMAIL",
                message=f"Invalid email format: {employee_email}",
                column="employee_email",
                value=employee_email
            ))

        raw_date = normalized.get("date")
        try:
            date_value = self._parse_date(raw_date)
        except ValueError as e:
            raise ParseIssueError(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.ERROR,
                code="INVALID_DATE",
                message=str(e),
                column="date",
                value=self._get_string(raw_date)
            ))
        if not date_value:
            raise ParseIssueError(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.ERROR,
                code="MISSING_REQUIRED_FIELD",
                message="Date is required",
                column="date"
            ))

        raw_start = normalized.get("start_time")
        try:
            start_time = self._parse_time(raw_start)
        except ValueError as e:
            raise ParseIssueError(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.ERROR,
                code="INVALID_TIME",
                message=str(e),
                column="start_time",
                value=self._get_string(raw_start)
            ))
        if not start_time:
            raise ParseIssueError(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.ERROR,
                code="MISSING_REQUIRED_FIELD",
                message="Start Time is required",
                column="start_time"
            ))

        raw_end = normalized.get("end_time")
        try:
            end_time = self._parse_time(raw_end)
        except ValueError as e:
            raise ParseIssueError(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.ERROR,
                code="INVALID_TIME",
                message=str(e),
                column="end_time",
                value=self._get_string(raw_end)
            ))
        if not end_time:
            raise ParseIssueError(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.ERROR,
                code="MISSING_REQUIRED_FIELD",
                message="End Time is required",
                column="end_time"
            ))

        # Detect overnight shift (end < start)
        is_overnight = end_time < start_time

        # Parse optional boolean fields with validation warnings
        has_meal_break = self._parse_boolean(normalized.get("has_meal_break"))
        meal_break_duration = self._parse_int(normalized.get("meal_break_duration"))

        # Warning: has_meal_break=True but no duration
        if has_meal_break and meal_break_duration is None:
            warnings.append(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.WARNING,
                code="MEAL_BREAK_DURATION_MISSING",
                message="has_meal_break=true but meal_break_duration is missing",
                column="meal_break_duration"
            ))

        has_rest_breaks = self._parse_boolean(normalized.get("has_rest_breaks"))
        rest_breaks_duration = self._parse_int(normalized.get("rest_breaks_duration"))

        # Warning: has_rest_breaks=True but no duration
        if has_rest_breaks and rest_breaks_duration is None:
            warnings.append(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.WARNING,
                code="REST_BREAKS_DURATION_MISSING",
                message="has_rest_breaks=true but rest_breaks_duration is missing",
                column="rest_breaks_duration"
            ))

        entry = RosterEntry(
            excel_row=row_num,
            employee_email=employee_email,
            employee_number=self._get_string(normalized.get("employee_number")),
            employee_name=self._get_string(normalized.get("employee_name")),
            date=date_value,
            start_time=start_time,
            end_time=end_time,
            is_overnight=is_overnight,
            has_meal_break=has_meal_break,
            meal_break_duration=meal_break_duration,
            has_rest_breaks=has_rest_breaks,
            rest_breaks_duration=rest_breaks_duration,
            is_public_holiday=self._parse_boolean(normalized.get("is_public_holiday")),
            public_holiday_name=self._get_string(normalized.get("public_holiday_name")),
            is_on_call=self._parse_boolean(normalized.get("is_on_call")),
            location=self._get_string(normalized.get("location")),
            notes=self._get_string(normalized.get("notes")),
        )

        return entry, warnings

    def _parse_employee_row(self, row: dict[str, Any], row_num: int) -> EmployeeEntry:
        """Parse a single row into an EmployeeEntry."""
        # Normalize headers to canonical keys
        normalized = self._normalize_headers(row)

        name = self._get_string(normalized.get("name"))
        if not name:
            raise ParseIssueError(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.ERROR,
                code="MISSING_REQUIRED_FIELD",
                message="Name is required",
                column="name"
            ))

        email = self._get_string(normalized.get("email"))
        if not email:
            raise ParseIssueError(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.ERROR,
                code="MISSING_REQUIRED_FIELD",
                message="Email is required",
                column="email"
            ))

        role = self._get_string(normalized.get("role"))
        if not role:
            raise ParseIssueError(ParseIssue(
                row=row_num,
                severity=ParseIssueSeverity.ERROR,
                code="MISSING_REQUIRED_FIELD",
                message="Role is required",
                column="role"
            ))

        return EmployeeEntry(
            excel_row=row_num,
            name=name,
            email=email,
            role=role,
            department=self._get_string(normalized.get("department")),
            start_date=self._parse_date(normalized.get("start_date")),
        )

    def _get_cell_value(self, cell: Cell) -> Any:
        """Extract value from an Excel cell, handling various types."""
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
            if value == "":
                return None
        return value

    def _get_string(self, value: Any) -> Optional[str]:
        """Convert value to string, return None if empty."""
        if value is None:
            return None
        result = str(value).strip()
        return result if result else None

    def _parse_date(self, value: Any) -> Optional[date]:
        """
        Parse various date formats into a date object.

        Supports:
        - datetime objects (from Excel)
        - date objects
        - Strings: YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY, DD-MM-YYYY
        """
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None

            # Try various formats
            formats = [
                "%Y-%m-%d",      # 2024-01-15
                "%d/%m/%Y",      # 15/01/2024
                "%m/%d/%Y",      # 01/15/2024
                "%d-%m-%Y",      # 15-01-2024
                "%Y/%m/%d",      # 2024/01/15
                "%d.%m.%Y",      # 15.01.2024
            ]

            for fmt in formats:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue

            raise ValueError(f"Unable to parse date: {value}")

        raise ValueError(f"Invalid date type: {type(value)}")

    def _parse_time(self, value: Any) -> Optional[time]:
        """
        Parse various time formats into a time object.

        Supports:
        - time objects
        - datetime objects (extracts time part)
        - Strings: HH:MM, HH:MM:SS, H:MM AM/PM
        - Float (Excel time as fraction of day)
        """
        if value is None:
            return None

        if isinstance(value, time):
            return value

        if isinstance(value, datetime):
            return value.time()

        if isinstance(value, (int, float)):
            # Excel stores time as fraction of day
            total_seconds = int(value * 24 * 60 * 60)
            hours = (total_seconds // 3600) % 24
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            return time(hours, minutes, seconds)

        if isinstance(value, str):
            value_stripped = value.strip()
            if not value_stripped:
                return None

            # Detect time ranges (e.g., "9:00-17:00", "9:00~17:00", "9:00 to 17:00")
            # These indicate user put both start and end time in one cell
            time_range_pattern = r'\d{1,2}[:\d]*\s*[-~～–—to]+\s*\d{1,2}'
            if re.search(time_range_pattern, value_stripped, re.IGNORECASE):
                raise ValueError(
                    f"Time range detected: '{value_stripped}'. "
                    "Please use separate 'Start Time' and 'End Time' columns."
                )

            value = value_stripped.upper()

            # Handle AM/PM format
            is_pm = "PM" in value
            is_am = "AM" in value
            value = value.replace("AM", "").replace("PM", "").strip()

            # Handle hour-only formats like "9", "9AM", "9 PM"
            if value.isdigit():
                hour = int(value)
                if is_pm and hour < 12:
                    hour += 12
                elif is_am and hour == 12:
                    hour = 0
                if 0 <= hour <= 23:
                    return time(hour, 0, 0)

            # Try various formats
            formats = [
                "%H:%M:%S",  # 14:30:00
                "%H:%M",     # 14:30
                "%I:%M:%S",  # 2:30:00 (12-hour)
                "%I:%M",     # 2:30 (12-hour)
            ]

            for fmt in formats:
                try:
                    parsed = datetime.strptime(value, fmt).time()
                    if is_pm and parsed.hour < 12:
                        parsed = time(parsed.hour + 12, parsed.minute, parsed.second)
                    elif is_am and parsed.hour == 12:
                        parsed = time(0, parsed.minute, parsed.second)
                    return parsed
                except ValueError:
                    continue

            raise ValueError(f"Unable to parse time: {value}")

        raise ValueError(f"Invalid time type: {type(value)}")

    def _parse_boolean(self, value: Any) -> bool:
        """
        Parse various boolean representations.

        Supports: True/False, Yes/No, Y/N, 1/0
        """
        if value is None:
            return False

        if isinstance(value, bool):
            return value

        if isinstance(value, (int, float)):
            return bool(value)

        if isinstance(value, str):
            value = value.strip().lower()
            return value in ("true", "yes", "y", "1", "on")

        return False

    def _parse_int(self, value: Any) -> Optional[int]:
        """Parse value as integer, return None if not valid."""
        if value is None:
            return None

        if isinstance(value, int):
            return value

        if isinstance(value, float):
            return int(value)

        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            try:
                return int(float(value))
            except ValueError:
                return None

        return None
