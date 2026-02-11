"""Excel file reader with header row detection."""

from __future__ import annotations

from typing import Any, Optional
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


def _get_cell_value(cell: Cell) -> Any:
    """Extract value from an Excel cell, handling various types."""
    value = cell.value
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
    return value


def read_excel(
    file_path: str,
    sheet_name: Optional[str] = None,
    header_row: Optional[int] = None,
) -> list[dict[str, Any]]:
    """
    Read an Excel file and return data as a list of dictionaries.

    Each dict includes '__row__' key with the real Excel row number.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Invalid file format. Expected .xlsx, got: {path.suffix}")

    try:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Failed to read Excel file: {str(exc)}") from exc

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    row_iter = sheet.iter_rows()

    header_cells = None
    if header_row is not None:
        header_index = max(header_row - 1, 0)
        for idx, row in enumerate(row_iter):
            if idx == header_index:
                header_cells = row
                break
    else:
        for row in row_iter:
            values = [_get_cell_value(cell) for cell in row]
            non_empty = [v for v in values if v not in (None, "")]
            if len(non_empty) >= 2:
                header_cells = row
                break

    if header_cells is None:
        if header_row is not None:
            workbook.close()
            raise ValueError(f"Header row not found: {header_row}")
        workbook.close()
        return []

    headers = [_get_cell_value(cell) for cell in header_cells]
    headers = [h.strip() if isinstance(h, str) else h for h in headers]

    data: list[dict[str, Any]] = []
    for row in row_iter:
        row_dict: dict[str, Any] = {}
        is_empty_row = True
        real_row_num: Optional[int] = None

        for cell in row:
            if hasattr(cell, "row") and cell.row is not None:
                real_row_num = cell.row
                break

        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                value = _get_cell_value(cell)
                if value is not None:
                    is_empty_row = False
                row_dict[headers[i]] = value

        if not is_empty_row and real_row_num is not None:
            row_dict["__row__"] = real_row_num
            data.append(row_dict)

    workbook.close()
    return data
