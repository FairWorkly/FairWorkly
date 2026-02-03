import pytest
from openpyxl import Workbook

from agents.roster.services.roster_import import ParseIssueSeverity, ParseMode
from agents.roster.services.roster_import.utils import parse_time

class TestOvernightShifts:
    """Tests for overnight shift detection."""

    def test_overnight_shift_detected(self, handler, temp_excel_path):
        """Test that overnight shifts are detected (end < start)."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Number", "Employee Email", "Date", "Start Time", "End Time"])
        ws.append(["EMP001", "john@example.com", "2024-01-15", "22:00", "06:00"])  # Overnight
        ws.append(["EMP002", "jane@example.com", "2024-01-15", "09:00", "17:00"])  # Normal
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result = response.result

        assert len(result.entries) == 2
        assert result.entries[0].is_overnight is True  # 22:00-06:00
        assert result.entries[1].is_overnight is False  # 09:00-17:00

    def test_midnight_end_not_overnight(self, handler, temp_excel_path):
        """Test that ending at midnight is not considered overnight."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Number", "Employee Email", "Date", "Start Time", "End Time"])
        ws.append(["EMP001", "john@example.com", "2024-01-15", "16:00", "00:00"])  # Ends at midnight
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result = response.result

        assert len(result.entries) == 1
        assert result.entries[0].is_overnight is True  # 00:00 < 16:00

class TestBreakWarnings:
    """Tests for break-related warnings."""

    def test_meal_break_without_duration_warning(self, handler, temp_excel_path):
        """Test warning when has_meal_break=True but duration is missing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Number", "Employee Email", "Employment Type", "Date", "Start Time", "End Time", "Has Meal Break", "Meal Break Duration"])
        ws.append(["EMP001", "john@example.com", "full-time", "2024-01-15", "09:00", "17:00", "Yes", None])
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result, issues = response.result, response.issues
        warnings = [issue for issue in issues if issue.severity == ParseIssueSeverity.WARNING.value]

        assert len(result.entries) == 1  # Entry is still created
        assert len(warnings) == 1
        assert warnings[0].severity == ParseIssueSeverity.WARNING.value
        assert warnings[0].column == "meal_break_duration"
        assert "has_meal_break=true" in warnings[0].message

    def test_rest_breaks_without_duration_warning(self, handler, temp_excel_path):
        """Test warning when has_rest_breaks=True but duration is missing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Number", "Employee Email", "Employment Type", "Date", "Start Time", "End Time", "Has Rest Breaks", "Rest Breaks Duration"])
        ws.append(["EMP001", "john@example.com", "full-time", "2024-01-15", "09:00", "17:00", "Yes", None])
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result, issues = response.result, response.issues
        warnings = [issue for issue in issues if issue.severity == ParseIssueSeverity.WARNING.value]

        assert len(result.entries) == 1
        assert len(warnings) == 1
        assert warnings[0].column == "rest_breaks_duration"

    def test_no_warning_when_break_false(self, handler, temp_excel_path):
        """Test no warning when has_meal_break=False and duration is missing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Number", "Employee Email", "Employment Type", "Date", "Start Time", "End Time", "Has Meal Break", "Meal Break Duration"])
        ws.append(["EMP001", "john@example.com", "full-time", "2024-01-15", "09:00", "17:00", "No", None])
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result, issues = response.result, response.issues
        warnings = [issue for issue in issues if issue.severity == ParseIssueSeverity.WARNING.value]

        assert len(result.entries) == 1
        assert len(warnings) == 0

class TestTimeRangeDetection:
    """Tests for time range detection in _parse_time()."""

    def test_hyphen_range_detected(self):
        """Test that 9:00-17:00 pattern is detected."""
        with pytest.raises(ValueError, match=r"Time range detected"):
            parse_time("9:00-17:00")

    def test_tilde_range_detected(self):
        """Test that 9:00~17:00 pattern is detected."""
        with pytest.raises(ValueError, match=r"Time range detected"):
            parse_time("9:00~17:00")

    def test_fullwidth_tilde_range_detected(self):
        """Test that 9:00～17:00 (full-width tilde) pattern is detected."""
        with pytest.raises(ValueError, match=r"Time range detected"):
            parse_time("9:00～17:00")

    def test_en_dash_range_detected(self):
        """Test that 9:00–17:00 (en-dash) pattern is detected."""
        with pytest.raises(ValueError, match=r"Time range detected"):
            parse_time("9:00–17:00")

    def test_em_dash_range_detected(self):
        """Test that 9:00—17:00 (em-dash) pattern is detected."""
        with pytest.raises(ValueError, match=r"Time range detected"):
            parse_time("9:00—17:00")

    def test_to_range_detected(self):
        """Test that '9:00 to 17:00' pattern is detected."""
        with pytest.raises(ValueError, match=r"Time range detected"):
            parse_time("9:00 to 17:00")

    def test_range_with_spaces_detected(self):
        """Test that '9:00 - 17:00' (with spaces) is detected."""
        with pytest.raises(ValueError, match=r"Time range detected"):
            parse_time("9:00 - 17:00")

    def test_simple_hour_range_detected(self):
        """Test that simple hour ranges like '9-17' are detected."""
        with pytest.raises(ValueError, match=r"Time range detected"):
            parse_time("9-17")

    def test_error_message_includes_original_value(self):
        """Test that error message includes the original value."""
        with pytest.raises(ValueError, match="9:00-17:00"):
            parse_time("9:00-17:00")

    def test_error_message_suggests_separate_columns(self):
        """Test that error message suggests using separate columns."""
        with pytest.raises(ValueError, match=r"separate.*Start Time.*End Time"):
            parse_time("9:00-17:00")

    def test_roster_with_time_range_returns_error(self, handler, temp_excel_path):
        """Test that roster parsing catches time range in cell."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Number", "Employee Email", "Date", "Start Time", "End Time"])
        ws.append(["EMP001", "john@example.com", "2024-01-15", "9:00-17:00", "17:00"])  # Range in Start Time
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result, issues = response.result, response.issues
        errors = [issue for issue in issues if issue.severity == ParseIssueSeverity.ERROR.value]

        assert len(result.entries) == 0
        assert len(errors) == 1
        assert "Time range detected" in errors[0].message
        assert errors[0].column == "start_time"
