import pytest
from openpyxl import Workbook, load_workbook

class TestReadExcel:
    """Tests for read_excel() function."""

    def test_read_valid_excel(self, handler, roster_excel):
        """Test reading a valid Excel file."""
        data = handler.read_excel(str(roster_excel))

        assert len(data) == 3
        assert data[0]["Employee Email"] == "john@example.com"
        assert data[0]["Employee Number"] == "EMP001"
        assert data[0]["Employee Name"] == "John Smith"
        assert data[0]["Location"] == "Office"

    def test_read_excel_includes_row_numbers(self, handler, roster_excel):
        """Test that read_excel includes real Excel row numbers."""
        data = handler.read_excel(str(roster_excel))

        assert "__row__" in data[0]
        assert data[0]["__row__"] == 2  # First data row is row 2 (after header)
        assert data[1]["__row__"] == 3
        assert data[2]["__row__"] == 4

    def test_read_excel_with_empty_rows(self, handler, temp_excel_path):
        """Test that empty rows are skipped but row numbers are preserved."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Value"])  # Row 1
        ws.append(["Test1", 100])      # Row 2
        ws.append([None, None])        # Row 3 - empty, should be skipped
        ws.append(["Test2", 200])      # Row 4
        wb.save(temp_excel_path)

        data = handler.read_excel(str(temp_excel_path))

        assert len(data) == 2
        assert data[0]["__row__"] == 2
        assert data[1]["__row__"] == 4  # Row 3 was skipped

    def test_read_nonexistent_file(self, handler):
        """Test reading a file that doesn't exist."""
        with pytest.raises(FileNotFoundError):
            handler.read_excel("/nonexistent/path/file.xlsx")

    def test_read_invalid_format(self, handler, tmp_path):
        """Test reading a file with invalid extension."""
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("not an excel file")

        with pytest.raises(ValueError, match=r"Invalid file format"):
            handler.read_excel(str(txt_file))

    def test_read_specific_sheet(self, handler, temp_excel_path):
        """Test reading a specific sheet by name."""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Col1"])
        ws1.append(["Value1"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Col2"])
        ws2.append(["Value2"])

        wb.save(temp_excel_path)

        # Read Sheet2
        data = handler.read_excel(str(temp_excel_path), sheet_name="Sheet2", header_row=1)
        assert len(data) == 1
        assert "Col2" in data[0]
        assert data[0]["Col2"] == "Value2"

    def test_read_invalid_sheet_name(self, handler, roster_excel):
        """Test reading a non-existent sheet."""
        with pytest.raises(ValueError, match=r"Sheet.*not found"):
            handler.read_excel(str(roster_excel), sheet_name="NonExistent")

    def test_read_invalid_sheet_name_closes_workbook(self, handler, temp_excel_path):
        """Invalid sheet name should not leave the workbook open."""
        wb = Workbook()
        wb.active.title = "Roster"
        wb.save(temp_excel_path)

        with pytest.raises(ValueError, match=r"Sheet.*not found"):
            handler.read_excel(str(temp_excel_path), sheet_name="Missing")

        reopened = load_workbook(temp_excel_path, read_only=True, data_only=True)
        reopened.close()

    def test_empty_file(self, handler, temp_excel_path):
        """Test reading an empty Excel file."""
        wb = Workbook()
        wb.save(temp_excel_path)

        data = handler.read_excel(str(temp_excel_path))
        assert data == []

    def test_handle_none_values(self, handler, temp_excel_path):
        """Test that None values are preserved."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Optional"])
        ws.append(["Test", None])
        wb.save(temp_excel_path)

        data = handler.read_excel(str(temp_excel_path))
        assert data[0]["Optional"] is None
