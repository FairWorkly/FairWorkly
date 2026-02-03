import asyncio
import pytest
from datetime import date, time
from decimal import Decimal
from openpyxl import Workbook

from agents.roster.services.roster_import import ParseIssueSeverity, ParseMode, ParseResultStatus

class TestValidateHeaders:
    """Tests for validate_headers() function."""

    def test_valid_headers(self, handler):
        """Test validation passes with all required headers."""
        data = [{"employee_number": "EMP001", "employee_email": "john@test.com", "date": "2024-01-15", "start_time": "09:00", "end_time": "17:00"}]
        is_valid, missing = handler.validate_headers(data, ["employee_number", "date"])

        assert is_valid is True
        assert missing == []

    def test_missing_headers(self, handler):
        """Test validation fails with missing headers."""
        data = [{"employee_number": "EMP001", "employee_email": "john@test.com"}]
        is_valid, missing = handler.validate_headers(data, ["employee_number", "date", "start_time"])

        assert is_valid is False
        assert "date" in missing
        assert "start_time" in missing

    def test_empty_data(self, handler):
        """Test validation with empty data."""
        is_valid, missing = handler.validate_headers([], ["employee_number"])

        assert is_valid is False
        assert "employee_number" in missing

class TestParseRosterExcel:
    """Tests for parse_roster_excel() function."""

    def test_parse_valid_roster(self, handler, roster_excel):
        """Test parsing a valid roster file."""
        response = handler.parse_roster_excel(str(roster_excel))

        result, issues = response.result, response.issues

        assert len(result.entries) == 3
        assert len(issues) == 0

        # Check first entry
        assert result.entries[0].employee_email == "john@example.com"
        assert result.entries[0].employee_number == "EMP001"
        assert result.entries[0].employee_name == "John Smith"
        assert result.entries[0].date == date(2024, 1, 15)
        assert result.entries[0].start_time == time(9, 0)
        assert result.entries[0].end_time == time(17, 0)
        assert result.entries[0].has_meal_break is True
        assert result.entries[0].meal_break_duration == 30
        assert result.entries[0].location == "Office"
        assert result.entries[0].excel_row == 2  # Real Excel row number

        # Check third entry
        assert result.entries[2].employee_email == "bob@example.com"
        assert result.entries[2].employee_number == "EMP003"
        assert result.entries[2].excel_row == 4

        # Summary checks
        assert result.week_start_date == date(2024, 1, 15)
        assert result.week_end_date == date(2024, 1, 16)
        assert result.total_shifts == 3
        assert result.total_hours == Decimal("24.00")
        assert result.unique_employees == 3

    def test_missing_required_columns(self, handler, temp_excel_path):
        """Test error when required columns are missing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Number", "Employee Email", "Date"])  # Missing Start Time, End Time
        ws.append(["EMP001", "john@example.com", "2024-01-15"])
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result, issues = response.result, response.issues
        assert len(result.entries) == 0
        assert len(issues) == 1
        assert issues[0].code == "MISSING_REQUIRED_COLUMNS"

    def test_row_level_errors(self, handler, temp_excel_path):
        """Test that row-level errors don't stop processing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Number", "Employee Email", "Date", "Start Time", "End Time"])
        ws.append(["EMP001", "john@example.com", "2024-01-15", "09:00", "17:00"])  # Valid
        ws.append(["EMP002", "jane@example.com", "invalid-date", "09:00", "17:00"])  # Invalid date
        ws.append(["EMP003", "bob@example.com", "2024-01-16", "08:00", "16:00"])  # Valid
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result, issues = response.result, response.issues
        errors = [issue for issue in issues if issue.severity == ParseIssueSeverity.ERROR.value]

        assert len(result.entries) == 2  # john and bob
        assert len(errors) == 1  # jane's row
        assert errors[0].row == 3  # Real Excel row number
        assert errors[0].severity == ParseIssueSeverity.ERROR.value

    def test_empty_roster_file(self, handler, temp_excel_path):
        """Test parsing an empty roster file."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Number", "Employee Email", "Date", "Start Time", "End Time"])
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result, issues = response.result, response.issues
        errors = [issue for issue in issues if issue.severity == ParseIssueSeverity.ERROR.value]

        assert len(result.entries) == 0
        assert len(errors) == 1
        assert "empty" in errors[0].message.lower()

    def test_invalid_email_format(self, handler, temp_excel_path):
        """Test that invalid email format is caught."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Number", "Employee Email", "Date", "Start Time", "End Time"])
        ws.append(["EMP001", "john@example.com", "2024-01-15", "09:00", "17:00"])  # Valid
        ws.append(["EMP002", "not-an-email", "2024-01-15", "09:00", "17:00"])  # Invalid email
        ws.append(["EMP003", "jane@example.com", "2024-01-16", "08:00", "16:00"])  # Valid
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result, issues = response.result, response.issues
        errors = [issue for issue in issues if issue.severity == ParseIssueSeverity.ERROR.value]

        assert len(result.entries) == 2  # john and jane
        assert len(errors) == 1  # invalid email row
        assert "Invalid email format" in errors[0].message


class TestParseExcel:
    """Tests for parse_excel() temp file handling."""

    def test_parse_excel_cleans_temp_on_read_error(self, handler, tmp_path, monkeypatch):
        """Ensure temp file is removed if file.read() raises."""
        temp_file = tmp_path / "upload.xlsx"

        class DummyTmpFile:
            def __init__(self, path):
                self.name = str(path)
                path.write_bytes(b"")

            def write(self, _data):
                return None

            def __enter__(self):
                return self

            def __exit__(self, _exc_type, _exc, _tb):
                return None

        def fake_named_tempfile(*_args, **_kwargs):
            return DummyTmpFile(temp_file)

        monkeypatch.setattr("tempfile.NamedTemporaryFile", fake_named_tempfile)

        class DummyUploadFile:
            filename = "upload.xlsx"

            async def read(self):
                raise RuntimeError("read failed")

        with pytest.raises(RuntimeError, match=r"read failed"):
            asyncio.run(handler.parse_excel(DummyUploadFile()))

        assert not temp_file.exists()
