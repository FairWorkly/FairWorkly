import pytest
from openpyxl import Workbook

from agents.roster.services.roster_import.header_map import HeaderMap, normalize_header

class TestHeaderAliases:
    """Tests for header alias mapping."""

    def test_snake_case_headers(self, handler, temp_excel_path):
        """Test that snake_case headers are recognized."""
        wb = Workbook()
        ws = wb.active
        ws.append(["employee_number", "employee_email", "date", "start_time", "end_time"])
        ws.append(["EMP001", "john@example.com", "2024-01-15", "09:00", "17:00"])
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result = response.result
        assert len(result.entries) == 1
        assert result.entries[0].employee_email == "john@example.com"

class TestNormalizeHeader:
    """Tests for normalize_header behavior."""

    def test_normalizes_whitespace_and_case(self):
        assert normalize_header(" Employee  Email ") == "employee email"
        assert normalize_header("EMPLOYEE\nEMAIL") == "employee email"

    def test_no_longer_normalizes_fullwidth_and_symbols(self):
        """Simplified normalization no longer handles full-width or special characters."""
        # Full-width characters are NOT converted (template-only mode)
        assert normalize_header("Ｅｍｐｌｏｙｅｅ＿Ｅｍａｉｌ") == "ｅｍｐｌｏｙｅｅ＿ｅｍａｉｌ"
        # Special characters are NOT removed (template uses standard names)
        assert normalize_header("Employee*Email") == "employee*email"
        assert normalize_header("Employee-Email") == "employee-email"

class TestDuplicateCanonicalDetection:
    """Tests for duplicate canonical header detection."""

    def test_duplicate_headers_detected(self):
        """Test duplicate detection with template column names."""
        mapper = HeaderMap()
        row = {
            "Employee Email": "john@example.com",
            "employee_email": "john.staff@example.com",  # Same canonical key
        }
        dupes = mapper.detect_duplicate_canonical_headers(row)
        assert "employee_email" in dupes
        assert len(dupes["employee_email"]) == 2

    def test_mixed_case_headers(self, handler, temp_excel_path):
        """Test that mixed case template headers are recognized."""
        wb = Workbook()
        ws = wb.active
        ws.append(["EMPLOYEE NUMBER", "EMPLOYEE EMAIL", "DATE", "START TIME", "END TIME"])
        ws.append(["EMP001", "john@example.com", "2024-01-15", "09:00", "17:00"])
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result = response.result
        assert len(result.entries) == 1
        assert result.entries[0].employee_email == "john@example.com"

    def test_whitespace_tolerance(self, handler, temp_excel_path):
        """Test that whitespace variations in template headers are tolerated."""
        wb = Workbook()
        ws = wb.active
        ws.append(["  Employee Number  ", "  Employee Email  ", "Date", " Start Time", "End Time "])
        ws.append(["EMP001", "john@example.com", "2024-01-15", "09:00", "17:00"])
        wb.save(temp_excel_path)

        response = handler.parse_roster_excel(str(temp_excel_path))

        result = response.result
        assert len(result.entries) == 1
        assert result.entries[0].employee_email == "john@example.com"
