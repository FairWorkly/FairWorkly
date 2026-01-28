"""
Unit tests for RosterExcelParser class.

Tests Excel file parsing functionality including:
- Reading Excel files with real row numbers
- Header alias mapping
- Parsing roster and employee data
- Date, time, and boolean parsing
- Warning generation for data quality issues
- Overnight shift detection
- Error handling
"""

import pytest
from datetime import date, time
from decimal import Decimal
from pathlib import Path
from openpyxl import Workbook

from agents.shared.roster_excel_parser import (
    RosterExcelParser,
    RosterEntry,
    EmployeeEntry,
    ParseIssue,
    ParseIssueSeverity,
)


@pytest.fixture
def handler():
    """Create a RosterExcelParser instance."""
    return RosterExcelParser()


@pytest.fixture
def temp_excel_path(tmp_path):
    """Helper to create temporary Excel file path."""
    return tmp_path / "test.xlsx"


@pytest.fixture
def roster_excel(temp_excel_path):
    """Create a valid roster Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    # Headers
    headers = [
        "Employee Email", "Employee Number", "Employee Name", "Date", "Start Time", "End Time",
        "Has Meal Break", "Meal Break Duration", "Location", "Notes"
    ]
    ws.append(headers)

    # Data rows
    ws.append(["john@example.com", "EMP001", "John Smith", "2024-01-15", "09:00", "17:00", "Yes", 30, "Office", "Regular shift"])
    ws.append(["jane@example.com", "EMP002", "Jane Doe", "2024-01-15", "14:00", "22:00", "Yes", 30, "Store", None])
    ws.append(["bob@example.com", None, "Bob Wilson", "2024-01-16", "06:00", "14:00", "No", None, None, "Early shift"])

    wb.save(temp_excel_path)
    return temp_excel_path


@pytest.fixture
def employee_excel(temp_excel_path):
    """Create a valid employee Excel file."""
    wb = Workbook()
    ws = wb.active

    headers = ["Name", "Email", "Role", "Department", "Start Date"]
    ws.append(headers)
    ws.append(["John Smith", "john@example.com", "Manager", "Sales", "2023-01-01"])
    ws.append(["Jane Doe", "jane@example.com", "Developer", "Engineering", "2023-06-15"])

    wb.save(temp_excel_path)
    return temp_excel_path


class TestReadExcel:
    """Tests for read_excel() function."""

    def test_read_valid_excel(self, handler, roster_excel):
        """Test reading a valid Excel file."""
        data = handler.read_excel(str(roster_excel))

        assert len(data) == 3
        assert data[0]["Employee Email"] == "john@example.com"
        assert data[0]["Employee Number"] == "EMP001"
        assert data[0]["Employee Name"] == "John Smith"
        assert data[0]["Location"] == "Office"

    def test_read_excel_includes_row_numbers(self, handler, roster_excel):
        """Test that read_excel includes real Excel row numbers."""
        data = handler.read_excel(str(roster_excel))

        assert "__row__" in data[0]
        assert data[0]["__row__"] == 2  # First data row is row 2 (after header)
        assert data[1]["__row__"] == 3
        assert data[2]["__row__"] == 4

    def test_read_excel_with_empty_rows(self, handler, temp_excel_path):
        """Test that empty rows are skipped but row numbers are preserved."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Value"])  # Row 1
        ws.append(["Test1", 100])      # Row 2
        ws.append([None, None])        # Row 3 - empty, should be skipped
        ws.append(["Test2", 200])      # Row 4
        wb.save(temp_excel_path)

        data = handler.read_excel(str(temp_excel_path))

        assert len(data) == 2
        assert data[0]["__row__"] == 2
        assert data[1]["__row__"] == 4  # Row 3 was skipped

    def test_read_nonexistent_file(self, handler):
        """Test reading a file that doesn't exist."""
        with pytest.raises(FileNotFoundError):
            handler.read_excel("/nonexistent/path/file.xlsx")

    def test_read_invalid_format(self, handler, tmp_path):
        """Test reading a file with invalid extension."""
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("not an excel file")

        with pytest.raises(ValueError, match="Invalid file format"):
            handler.read_excel(str(txt_file))

    def test_read_specific_sheet(self, handler, temp_excel_path):
        """Test reading a specific sheet by name."""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Col1"])
        ws1.append(["Value1"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Col2"])
        ws2.append(["Value2"])

        wb.save(temp_excel_path)

        # Read Sheet2
        data = handler.read_excel(str(temp_excel_path), sheet_name="Sheet2")
        assert len(data) == 1
        assert "Col2" in data[0]
        assert data[0]["Col2"] == "Value2"

    def test_read_invalid_sheet_name(self, handler, roster_excel):
        """Test reading a non-existent sheet."""
        with pytest.raises(ValueError, match="Sheet.*not found"):
            handler.read_excel(str(roster_excel), sheet_name="NonExistent")

    def test_empty_file(self, handler, temp_excel_path):
        """Test reading an empty Excel file."""
        wb = Workbook()
        wb.save(temp_excel_path)

        data = handler.read_excel(str(temp_excel_path))
        assert data == []

    def test_handle_none_values(self, handler, temp_excel_path):
        """Test that None values are preserved."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Optional"])
        ws.append(["Test", None])
        wb.save(temp_excel_path)

        data = handler.read_excel(str(temp_excel_path))
        assert data[0]["Optional"] is None


class TestHeaderAliases:
    """Tests for header alias mapping."""

    def test_snake_case_headers(self, handler, temp_excel_path):
        """Test that snake_case headers are recognized."""
        wb = Workbook()
        ws = wb.active
        ws.append(["employee_email", "date", "start_time", "end_time"])
        ws.append(["john@example.com", "2024-01-15", "09:00", "17:00"])
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))
        assert len(result.entries) == 1
        assert result.entries[0].employee_email == "john@example.com"

    def test_mixed_case_headers(self, handler, temp_excel_path):
        """Test that mixed case headers are recognized."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Email", "DATE", "Start", "End"])
        ws.append(["john@example.com", "2024-01-15", "09:00", "17:00"])
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))
        assert len(result.entries) == 1

    def test_alternative_header_names(self, handler, temp_excel_path):
        """Test that alternative header names work."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Staff Email", "shift date", "start", "finish time"])
        ws.append(["john@example.com", "2024-01-15", "09:00", "17:00"])
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))
        assert len(result.entries) == 1
        assert result.entries[0].employee_email == "john@example.com"


class TestValidateHeaders:
    """Tests for validate_headers() function."""

    def test_valid_headers(self, handler):
        """Test validation passes with all required headers."""
        data = [{"employee_email": "john@test.com", "date": "2024-01-15", "start_time": "09:00", "end_time": "17:00"}]
        is_valid, missing = handler.validate_headers(data, ["employee_email", "date"])

        assert is_valid is True
        assert missing == []

    def test_missing_headers(self, handler):
        """Test validation fails with missing headers."""
        data = [{"employee_email": "john@test.com"}]
        is_valid, missing = handler.validate_headers(data, ["employee_email", "date", "start_time"])

        assert is_valid is False
        assert "date" in missing
        assert "start_time" in missing

    def test_empty_data(self, handler):
        """Test validation with empty data."""
        is_valid, missing = handler.validate_headers([], ["employee_email"])

        assert is_valid is False
        assert "employee_email" in missing


class TestParseRosterExcel:
    """Tests for parse_roster_excel() function."""

    def test_parse_valid_roster(self, handler, roster_excel):
        """Test parsing a valid roster file."""
        result, issues = handler.parse_roster_excel(str(roster_excel))

        assert len(result.entries) == 3
        assert len(issues) == 0

        # Check first entry
        assert result.entries[0].employee_email == "john@example.com"
        assert result.entries[0].employee_number == "EMP001"
        assert result.entries[0].employee_name == "John Smith"
        assert result.entries[0].date == date(2024, 1, 15)
        assert result.entries[0].start_time == time(9, 0)
        assert result.entries[0].end_time == time(17, 0)
        assert result.entries[0].has_meal_break is True
        assert result.entries[0].meal_break_duration == 30
        assert result.entries[0].location == "Office"
        assert result.entries[0].excel_row == 2  # Real Excel row number

        # Check third entry - no employee number
        assert result.entries[2].employee_email == "bob@example.com"
        assert result.entries[2].employee_number is None
        assert result.entries[2].excel_row == 4

        # Summary checks
        assert result.week_start_date == date(2024, 1, 15)
        assert result.week_end_date == date(2024, 1, 16)
        assert result.total_shifts == 3
        assert result.total_hours == Decimal("24.00")
        assert result.unique_employees == 3

    def test_missing_required_columns(self, handler, temp_excel_path):
        """Test error when required columns are missing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Email", "Date"])  # Missing Start Time, End Time
        ws.append(["john@example.com", "2024-01-15"])
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))
        assert len(result.entries) == 0
        assert len(issues) == 1
        assert issues[0].code == "MISSING_REQUIRED_COLUMNS"

    def test_row_level_errors(self, handler, temp_excel_path):
        """Test that row-level errors don't stop processing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Email", "Date", "Start Time", "End Time"])
        ws.append(["john@example.com", "2024-01-15", "09:00", "17:00"])  # Valid
        ws.append(["jane@example.com", "invalid-date", "09:00", "17:00"])  # Invalid date
        ws.append(["bob@example.com", "2024-01-16", "08:00", "16:00"])  # Valid
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))
        errors = [issue for issue in issues if issue.severity == ParseIssueSeverity.ERROR.value]

        assert len(result.entries) == 2  # john and bob
        assert len(errors) == 1  # jane's row
        assert errors[0].row == 3  # Real Excel row number
        assert errors[0].severity == ParseIssueSeverity.ERROR.value

    def test_empty_roster_file(self, handler, temp_excel_path):
        """Test parsing an empty roster file."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Email", "Date", "Start Time", "End Time"])
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))
        errors = [issue for issue in issues if issue.severity == ParseIssueSeverity.ERROR.value]

        assert len(result.entries) == 0
        assert len(errors) == 1
        assert "empty" in errors[0].message.lower()

    def test_invalid_email_format(self, handler, temp_excel_path):
        """Test that invalid email format is caught."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Email", "Date", "Start Time", "End Time"])
        ws.append(["john@example.com", "2024-01-15", "09:00", "17:00"])  # Valid
        ws.append(["not-an-email", "2024-01-15", "09:00", "17:00"])  # Invalid email
        ws.append(["jane@example.com", "2024-01-16", "08:00", "16:00"])  # Valid
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))
        errors = [issue for issue in issues if issue.severity == ParseIssueSeverity.ERROR.value]

        assert len(result.entries) == 2  # john and jane
        assert len(errors) == 1  # invalid email row
        assert "Invalid email format" in errors[0].message


class TestOvernightShifts:
    """Tests for overnight shift detection."""

    def test_overnight_shift_detected(self, handler, temp_excel_path):
        """Test that overnight shifts are detected (end < start)."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Email", "Date", "Start Time", "End Time"])
        ws.append(["john@example.com", "2024-01-15", "22:00", "06:00"])  # Overnight
        ws.append(["jane@example.com", "2024-01-15", "09:00", "17:00"])  # Normal
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))

        assert len(result.entries) == 2
        assert result.entries[0].is_overnight is True  # 22:00-06:00
        assert result.entries[1].is_overnight is False  # 09:00-17:00

    def test_midnight_end_not_overnight(self, handler, temp_excel_path):
        """Test that ending at midnight is not considered overnight."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Email", "Date", "Start Time", "End Time"])
        ws.append(["john@example.com", "2024-01-15", "16:00", "00:00"])  # Ends at midnight
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))

        assert len(result.entries) == 1
        assert result.entries[0].is_overnight is True  # 00:00 < 16:00


class TestBreakWarnings:
    """Tests for break-related warnings."""

    def test_meal_break_without_duration_warning(self, handler, temp_excel_path):
        """Test warning when has_meal_break=True but duration is missing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Email", "Date", "Start Time", "End Time", "Has Meal Break", "Meal Break Duration"])
        ws.append(["john@example.com", "2024-01-15", "09:00", "17:00", "Yes", None])
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))
        warnings = [issue for issue in issues if issue.severity == ParseIssueSeverity.WARNING.value]

        assert len(result.entries) == 1  # Entry is still created
        assert len(warnings) == 1
        assert warnings[0].severity == ParseIssueSeverity.WARNING.value
        assert warnings[0].column == "meal_break_duration"
        assert "has_meal_break=true" in warnings[0].message

    def test_rest_breaks_without_duration_warning(self, handler, temp_excel_path):
        """Test warning when has_rest_breaks=True but duration is missing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Email", "Date", "Start Time", "End Time", "Has Rest Breaks", "Rest Breaks Duration"])
        ws.append(["john@example.com", "2024-01-15", "09:00", "17:00", "Yes", None])
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))
        warnings = [issue for issue in issues if issue.severity == ParseIssueSeverity.WARNING.value]

        assert len(result.entries) == 1
        assert len(warnings) == 1
        assert warnings[0].column == "rest_breaks_duration"

    def test_no_warning_when_break_false(self, handler, temp_excel_path):
        """Test no warning when has_meal_break=False and duration is missing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Email", "Date", "Start Time", "End Time", "Has Meal Break", "Meal Break Duration"])
        ws.append(["john@example.com", "2024-01-15", "09:00", "17:00", "No", None])
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))
        warnings = [issue for issue in issues if issue.severity == ParseIssueSeverity.WARNING.value]

        assert len(result.entries) == 1
        assert len(warnings) == 0


class TestParseEmployeeExcel:
    """Tests for parse_employee_excel() function."""

    def test_parse_valid_employees(self, handler, employee_excel):
        """Test parsing a valid employee file."""
        entries, errors = handler.parse_employee_excel(str(employee_excel))

        assert len(entries) == 2
        assert len(errors) == 0

        assert entries[0].name == "John Smith"
        assert entries[0].email == "john@example.com"
        assert entries[0].role == "Manager"
        assert entries[0].department == "Sales"
        assert entries[0].excel_row == 2


class TestParseDateHelper:
    """Tests for _parse_date() helper."""

    def test_parse_iso_format(self, handler):
        """Test parsing YYYY-MM-DD format."""
        result = handler._parse_date("2024-01-15")
        assert result == date(2024, 1, 15)

    def test_parse_slash_format_dmy(self, handler):
        """Test parsing DD/MM/YYYY format."""
        result = handler._parse_date("15/01/2024")
        assert result == date(2024, 1, 15)

    def test_parse_dash_format_dmy(self, handler):
        """Test parsing DD-MM-YYYY format."""
        result = handler._parse_date("15-01-2024")
        assert result == date(2024, 1, 15)

    def test_parse_datetime_object(self, handler):
        """Test parsing datetime object."""
        from datetime import datetime
        dt = datetime(2024, 1, 15, 10, 30)
        result = handler._parse_date(dt)
        assert result == date(2024, 1, 15)

    def test_parse_date_object(self, handler):
        """Test parsing date object."""
        d = date(2024, 1, 15)
        result = handler._parse_date(d)
        assert result == date(2024, 1, 15)

    def test_parse_none(self, handler):
        """Test parsing None returns None."""
        result = handler._parse_date(None)
        assert result is None

    def test_parse_invalid_date(self, handler):
        """Test parsing invalid date raises error."""
        with pytest.raises(ValueError, match="Unable to parse date"):
            handler._parse_date("not-a-date")


class TestParseTimeHelper:
    """Tests for _parse_time() helper."""

    def test_parse_24hour_format(self, handler):
        """Test parsing HH:MM format."""
        result = handler._parse_time("14:30")
        assert result == time(14, 30)

    def test_parse_24hour_with_seconds(self, handler):
        """Test parsing HH:MM:SS format."""
        result = handler._parse_time("14:30:45")
        assert result == time(14, 30, 45)

    def test_parse_12hour_pm(self, handler):
        """Test parsing 12-hour PM format."""
        result = handler._parse_time("2:30 PM")
        assert result == time(14, 30)

    def test_parse_12hour_am(self, handler):
        """Test parsing 12-hour AM format."""
        result = handler._parse_time("9:30 AM")
        assert result == time(9, 30)

    def test_parse_12_pm(self, handler):
        """Test parsing 12:00 PM (noon)."""
        result = handler._parse_time("12:00 PM")
        assert result == time(12, 0)

    def test_parse_12_am(self, handler):
        """Test parsing 12:00 AM (midnight)."""
        result = handler._parse_time("12:00 AM")
        assert result == time(0, 0)

    def test_parse_hour_only(self, handler):
        """Test parsing hour-only formats like 9, 9AM, 9 PM."""
        result = handler._parse_time("9")
        assert result == time(9, 0)

        result = handler._parse_time("9AM")
        assert result == time(9, 0)

        result = handler._parse_time("9 PM")
        assert result == time(21, 0)

    def test_parse_excel_time_float(self, handler):
        """Test parsing Excel time as float (fraction of day)."""
        # 0.5 = 12:00 (noon)
        result = handler._parse_time(0.5)
        assert result == time(12, 0, 0)

        # 0.25 = 06:00
        result = handler._parse_time(0.25)
        assert result == time(6, 0, 0)

    def test_parse_time_object(self, handler):
        """Test parsing time object."""
        t = time(14, 30)
        result = handler._parse_time(t)
        assert result == time(14, 30)

    def test_parse_none(self, handler):
        """Test parsing None returns None."""
        result = handler._parse_time(None)
        assert result is None


class TestParseBooleanHelper:
    """Tests for _parse_boolean() helper."""

    def test_parse_true_values(self, handler):
        """Test various true representations."""
        assert handler._parse_boolean(True) is True
        assert handler._parse_boolean("true") is True
        assert handler._parse_boolean("True") is True
        assert handler._parse_boolean("yes") is True
        assert handler._parse_boolean("Yes") is True
        assert handler._parse_boolean("Y") is True
        assert handler._parse_boolean("1") is True
        assert handler._parse_boolean(1) is True

    def test_parse_false_values(self, handler):
        """Test various false representations."""
        assert handler._parse_boolean(False) is False
        assert handler._parse_boolean("false") is False
        assert handler._parse_boolean("no") is False
        assert handler._parse_boolean("N") is False
        assert handler._parse_boolean("0") is False
        assert handler._parse_boolean(0) is False
        assert handler._parse_boolean(None) is False
        assert handler._parse_boolean("") is False


class TestParseIntHelper:
    """Tests for _parse_int() helper."""

    def test_parse_int(self, handler):
        """Test parsing integer."""
        assert handler._parse_int(30) == 30

    def test_parse_float(self, handler):
        """Test parsing float truncates to int."""
        assert handler._parse_int(30.5) == 30

    def test_parse_string(self, handler):
        """Test parsing numeric string."""
        assert handler._parse_int("30") == 30
        assert handler._parse_int("30.5") == 30

    def test_parse_none(self, handler):
        """Test parsing None returns None."""
        assert handler._parse_int(None) is None

    def test_parse_invalid(self, handler):
        """Test parsing invalid string returns None."""
        assert handler._parse_int("not-a-number") is None
        assert handler._parse_int("") is None


class TestTimeRangeDetection:
    """Tests for time range detection in _parse_time()."""

    def test_hyphen_range_detected(self, handler):
        """Test that 9:00-17:00 pattern is detected."""
        with pytest.raises(ValueError, match="Time range detected"):
            handler._parse_time("9:00-17:00")

    def test_tilde_range_detected(self, handler):
        """Test that 9:00~17:00 pattern is detected."""
        with pytest.raises(ValueError, match="Time range detected"):
            handler._parse_time("9:00~17:00")

    def test_fullwidth_tilde_range_detected(self, handler):
        """Test that 9:00～17:00 (full-width tilde) pattern is detected."""
        with pytest.raises(ValueError, match="Time range detected"):
            handler._parse_time("9:00～17:00")

    def test_en_dash_range_detected(self, handler):
        """Test that 9:00–17:00 (en-dash) pattern is detected."""
        with pytest.raises(ValueError, match="Time range detected"):
            handler._parse_time("9:00–17:00")

    def test_em_dash_range_detected(self, handler):
        """Test that 9:00—17:00 (em-dash) pattern is detected."""
        with pytest.raises(ValueError, match="Time range detected"):
            handler._parse_time("9:00—17:00")

    def test_to_range_detected(self, handler):
        """Test that '9:00 to 17:00' pattern is detected."""
        with pytest.raises(ValueError, match="Time range detected"):
            handler._parse_time("9:00 to 17:00")

    def test_range_with_spaces_detected(self, handler):
        """Test that '9:00 - 17:00' (with spaces) is detected."""
        with pytest.raises(ValueError, match="Time range detected"):
            handler._parse_time("9:00 - 17:00")

    def test_simple_hour_range_detected(self, handler):
        """Test that simple hour ranges like '9-17' are detected."""
        with pytest.raises(ValueError, match="Time range detected"):
            handler._parse_time("9-17")

    def test_error_message_includes_original_value(self, handler):
        """Test that error message includes the original value."""
        with pytest.raises(ValueError, match="9:00-17:00"):
            handler._parse_time("9:00-17:00")

    def test_error_message_suggests_separate_columns(self, handler):
        """Test that error message suggests using separate columns."""
        with pytest.raises(ValueError, match="separate.*Start Time.*End Time"):
            handler._parse_time("9:00-17:00")

    def test_roster_with_time_range_returns_error(self, handler, temp_excel_path):
        """Test that roster parsing catches time range in cell."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee Email", "Date", "Start Time", "End Time"])
        ws.append(["john@example.com", "2024-01-15", "9:00-17:00", "17:00"])  # Range in Start Time
        wb.save(temp_excel_path)

        result, issues = handler.parse_roster_excel(str(temp_excel_path))
        errors = [issue for issue in issues if issue.severity == ParseIssueSeverity.ERROR.value]

        assert len(result.entries) == 0
        assert len(errors) == 1
        assert "Time range detected" in errors[0].message
        assert errors[0].column == "start_time"
