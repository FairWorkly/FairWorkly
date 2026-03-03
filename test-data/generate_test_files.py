#!/usr/bin/env python3
"""Generate test Excel files for roster compliance testing.

Usage: python test-data/generate_test_files.py
Files are saved into the test-data/ directory alongside this script.
"""

from pathlib import Path
from openpyxl import Workbook
from datetime import date, time, timedelta

OUTPUT_DIR = Path(__file__).parent


def create_parser_errors_file():
    """Test file 1: Parser error scenarios."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    # Headers
    headers = [
        "Employee Number",
        "Employee Name",
        "Date",
        "Start Time",
        "End Time",
        "Has Meal Break",
        "Meal Break Duration",
        "Employment Type",
        "Notes",
    ]
    ws.append(headers)

    # Row 1: Missing employee number (MISSING_REQUIRED_FIELD)
    ws.append(["", "Missing EmpNo", "2026-02-03", "09:00", "17:00", "No", "", "Casual", "Should fail: missing employee number"])

    # Row 2: Invalid date format (INVALID_DATE)
    ws.append(["EMP001", "Sarah Chen", "invalid-date", "09:00", "17:00", "No", "", "Casual", "Should fail: invalid date"])

    # Row 3: Invalid time format (INVALID_TIME)
    ws.append(["EMP002", "Michael Nguyen", "2026-02-03", "25:00", "17:00", "No", "", "Casual", "Should fail: invalid time"])

    # Row 4: Unrecognized employment type (EMPLOYMENT_TYPE_UNRECOGNIZED warning)
    ws.append(["EMP003", "Emma Wilson", "2026-02-03", "09:00", "17:00", "No", "", "Unknown", "Should warn: unknown employment type"])

    # Row 5: Overnight shift without explicit flag (OVERNIGHT_ASSUMED warning)
    ws.append(["EMP004", "James Taylor", "2026-02-03", "22:00", "06:00", "No", "", "FullTime", "Should warn: overnight assumed"])

    # Row 6: Valid row for comparison
    ws.append(["EMP001", "Sarah Chen", "2026-02-04", "09:00", "17:00", "Yes", "30", "FullTime", "Valid row"])

    wb.save(OUTPUT_DIR / "test-parser-errors.xlsx")
    print("Created: test-data/test-parser-errors.xlsx")


def create_compliance_errors_file():
    """Test file 2: Compliance rule violation scenarios."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    headers = [
        "Employee Number",
        "Employee Name",
        "Date",
        "Start Time",
        "End Time",
        "Has Meal Break",
        "Meal Break Duration",
        "Employment Type",
        "Notes",
    ]
    ws.append(headers)

    # Scenario 1: MealBreak Error - 6 hour shift without break
    ws.append(["EMP003", "Emma Wilson", "2026-02-03", "09:00", "15:00", "No", "", "Casual", "MealBreak Error: 6hr no break"])

    # Scenario 2: MinimumShiftHours Error - 2 hour shift for casual
    ws.append(["EMP003", "Emma Wilson", "2026-02-04", "09:00", "11:00", "No", "", "Casual", "MinShift Error: 2hr < 3hr min"])

    # Scenario 3: RestPeriod Warning - less than 10 hours between shifts
    ws.append(["EMP001", "Sarah Chen", "2026-02-03", "09:00", "18:00", "Yes", "30", "FullTime", "Shift 1 for rest period test"])
    ws.append(["EMP001", "Sarah Chen", "2026-02-04", "02:00", "10:00", "Yes", "30", "FullTime", "RestPeriod Warning: only 8hr rest"])

    # Scenario 4: ConsecutiveDays Warning - 7 consecutive days
    base_date = date(2026, 2, 9)
    for i in range(7):
        d = base_date + timedelta(days=i)
        ws.append([
            "EMP002",
            "Michael Nguyen",
            d.strftime("%Y-%m-%d"),
            "10:00",
            "14:00",
            "No",
            "",
            "PartTime",
            f"Day {i+1} of 7 consecutive" if i < 6 else "ConsecutiveDays Warning: 7 days"
        ])

    # Scenario 5: WeeklyHoursLimit Info - Full-time > 38 hours
    # Monday to Friday, 9 hours each = 45 hours
    week_start = date(2026, 2, 16)  # Monday
    for i in range(5):
        d = week_start + timedelta(days=i)
        ws.append([
            "EMP001",
            "Sarah Chen",
            d.strftime("%Y-%m-%d"),
            "08:00",
            "17:30",
            "Yes",
            "30",
            "FullTime",
            f"WeeklyHours test: day {i+1}, 9hr" if i < 4 else "WeeklyHours Info: 45hr > 38hr"
        ])

    # Scenario 6: DataQuality Warning - break exceeds shift
    ws.append(["EMP004", "James Taylor", "2026-02-22", "09:00", "11:00", "Yes", "180", "Casual", "DataQuality Warning: 180min break > 2hr shift"])

    wb.save(OUTPUT_DIR / "test-compliance-errors.xlsx")
    print("Created: test-data/test-compliance-errors.xlsx")


def create_happy_path_file():
    """Test file 3: Valid data with no errors."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    headers = [
        "Employee Number",
        "Employee Name",
        "Date",
        "Start Time",
        "End Time",
        "Has Meal Break",
        "Meal Break Duration",
        "Employment Type",
        "Notes",
    ]
    ws.append(headers)

    # Valid full-time shift with break
    ws.append(["EMP001", "Sarah Chen", "2026-02-03", "09:00", "17:00", "Yes", "30", "FullTime", "Valid 8hr shift"])

    # Valid part-time shift (4 hours, no break required)
    ws.append(["EMP002", "Michael Nguyen", "2026-02-03", "10:00", "14:00", "No", "", "PartTime", "Valid 4hr shift"])

    # Valid casual shift (4 hours)
    ws.append(["EMP003", "Emma Wilson", "2026-02-03", "18:00", "22:00", "No", "", "Casual", "Valid 4hr shift"])

    # Another valid casual shift (3 hours minimum)
    ws.append(["EMP004", "James Taylor", "2026-02-03", "09:00", "12:00", "No", "", "Casual", "Valid 3hr minimum"])

    # Valid week with proper rest periods
    base_date = date(2026, 2, 9)
    for i in range(3):
        d = base_date + timedelta(days=i * 2)  # Every other day = proper rest
        ws.append([
            "EMP001",
            "Sarah Chen",
            d.strftime("%Y-%m-%d"),
            "09:00",
            "17:00",
            "Yes",
            "30",
            "FullTime",
            f"Valid shift with proper rest"
        ])

    wb.save(OUTPUT_DIR / "test-happy-path.xlsx")
    print("Created: test-data/test-happy-path.xlsx")


if __name__ == "__main__":
    create_parser_errors_file()
    create_compliance_errors_file()
    create_happy_path_file()
    print("\nAll test files created successfully!")
